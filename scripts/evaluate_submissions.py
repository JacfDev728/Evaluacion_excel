# scripts/evaluate_submissions.py

import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# --- Configuración de rutas ---
import os

# --- Configuración de rutas ---
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.dirname(SCRIPT_DIR)
DATA_DIR = os.path.join(ROOT_DIR, 'data')
SUBMISSIONS_DIR = os.path.join(ROOT_DIR, 'user_submissions')

# Archivos de referencia
TEMPLATE_FILE = os.path.join(DATA_DIR, 'base_datos_original.xlsx')
EXPECTED_FILE = os.path.join(DATA_DIR, 'respuestas_esperadas.xlsx')
RESULTS_FILE = 'evaluation_results.xlsx'

# --- Variables globales para el informe ---
results = []

# --- Configuración de las preguntas ---
# Las preguntas se evalúan directamente en el código, no necesitamos una lista separada

# --- Configuración de columnas ---
COLUMN_MAPPING = {
    "ID": "C",
    "Nombre del Cliente": "D",
    "Sentimiento": "E",  # Originalmente "Seguimiento", ahora "Sentimiento"
    "Puntuación": "F",
    "Fecha": "G",
    "Motivo": "H",
    "Ciudad": "I",
    "Canal": "J",
    "Duración Llamada (Minutos)": "K"
}

# --- Configuración de filas ---
HEADER_ROW = 5  # La fila donde están los encabezados
DATA_START_ROW = 6  # La fila donde empiezan los datos

# --- Configuración de hojas ---
SHEET_NAME = 'Datos'  # Nombre de la hoja con los datos
ANSWERS_SHEET = 'Respuestas'  # Nombre de la hoja donde el usuario escribe sus respuestas


def add_result(question_num, topic, question_text, status, observations=""):
    """Añade un resultado a la lista global."""
    results.append({
        "No.": question_num,
        "Tema": topic,
        "Pregunta": question_text,
        "Estado": status,
        "Observaciones": observations
    })


def evaluate_submission(submission_path):
    """Evalúa un archivo de envío de usuario."""
    user_filename = os.path.basename(submission_path)
    add_result("", "", f"--- Evaluando: {user_filename} ---", "")

    try:
        # Verificar que el archivo existe
        if not os.path.exists(submission_path):
            add_result("", "", f"Error al procesar {user_filename}", "Error", "Archivo no encontrado")
            return

        # Cargar el archivo del usuario
        wb_user = load_workbook(submission_path)
        ws_user = wb_user.active  # Usar la hoja activa

        # Cargar el archivo de respuestas esperadas
        wb_expected = load_workbook(EXPECTED_FILE)
        ws_expected = wb_expected.active  # Usar la hoja activa

        # Verificar que las columnas están en las posiciones correctas
        for col_name, col_letter in COLUMN_MAPPING.items():
            cell_user = ws_user[f'{col_letter}{HEADER_ROW}'].value
            cell_expected = ws_expected[f'{col_letter}{HEADER_ROW}'].value
            
            if str(cell_user).strip() != col_name:
                add_result("", "", f"Error al procesar {user_filename}", "Error", 
                          f"Columna '{col_name}' no encontrada en la posición correcta")
                return

        # Verificar y limpiar los nombres de las columnas
        # Obtener los nombres de las columnas desde la hoja
        column_names = []
        for col_letter in COLUMN_MAPPING.values():
            cell_value = ws_user[f'{col_letter}{HEADER_ROW}'].value
            if cell_value:
                # Limpiar espacios y caracteres especiales
                clean_value = str(cell_value).strip()
                column_names.append(clean_value)
            else:
                column_names.append(None)

        # Verificar que todas las columnas requeridas existen
        missing_columns = []
        for col_name, col_letter in COLUMN_MAPPING.items():
            cell_value = column_names[list(COLUMN_MAPPING.values()).index(col_letter)]
            if cell_value != col_name:
                missing_columns.append(col_name)
                
        if missing_columns:
            # Mostrar los nombres reales de las columnas para depuración
            actual_cols = {k: v for k, v in zip(COLUMN_MAPPING.keys(), column_names)}
            add_result("", "", f"Error al procesar {user_filename}", "Error", 
                      f"Columnas faltantes o con nombres incorrectos: {', '.join(missing_columns)}.\n"
                      f"Columnas encontradas: {actual_cols}")
            return

        # Cargar el archivo esperado como referencia
        wb_expected = load_workbook(EXPECTED_FILE)
        
        # Verificar que la hoja 'Datos' existe en el archivo de respuestas esperadas
        sheet_names_expected = [name.strip() for name in wb_expected.sheetnames]
        if SHEET_NAME not in sheet_names_expected:
            add_result("", "", f"Error al procesar {user_filename}", "Error", 
                      f"Hoja '{SHEET_NAME}' no encontrada en el archivo de respuestas esperadas. Hojas disponibles: {', '.join(wb_expected.sheetnames)}")
            return
            
        # Encontrar la hoja correcta en el archivo de respuestas esperadas
        ws_expected = next(ws for ws in wb_expected.worksheets if ws.title.strip() == SHEET_NAME)

        # Cargar los datos usando pandas
        try:
            # Obtener el nombre exacto de la hoja para pandas
            sheet_name_user = wb_user.sheetnames[0]  # Usar la primera hoja encontrada
            sheet_name_expected = wb_expected.sheetnames[0]  # Usar la primera hoja encontrada

            # Generar el rango de columnas
            first_col_letter = COLUMN_MAPPING["ID"]
            last_col_letter = COLUMN_MAPPING["Duración Llamada (Minutos)"]
            column_range = f"{first_col_letter}:{last_col_letter}"

            # Cargar DataFrames usando el nombre exacto de la hoja
            df_user = pd.read_excel(submission_path, sheet_name=sheet_name_user, 
                                   header=HEADER_ROW - 1, usecols=column_range)
            
            # Limpiar los nombres de las columnas
            df_user.columns = df_user.columns.str.strip()
            
            df_expected = pd.read_excel(EXPECTED_FILE, sheet_name=sheet_name_expected, 
                                      header=HEADER_ROW - 1, usecols=column_range)
        except Exception as e:
            add_result("", "", f"Error al procesar {user_filename}", "Error", 
                      f"Error al cargar los datos: {str(e)}")
            return

        # Verificar que los DataFrames tienen las columnas esperadas
        expected_columns = list(COLUMN_MAPPING.keys())
        missing_df_columns = [col for col in expected_columns if col not in df_user.columns]
        
        if missing_df_columns:
            missing_cols_str = ", ".join(missing_df_columns)
            add_result("", "", f"Error al procesar {user_filename}", "Error", 
                      f"Columnas faltantes en DataFrame: {missing_cols_str}")
            return

        # --- PREGUNTAS DE EVALUACIÓN --
        #PREGUNTA #1
        try:
            # Buscar la respuesta en M6
            answer_cell = "M6"  # Celda específica para la pregunta 1
            user_answer = ws_user[answer_cell].value
            
            if user_answer is None:
                add_result(1, "Edición y formato", "Cuantos ID tiene la base de datos", "Error", "Respuesta no encontrada")
            else:
                # Convertir a número si es necesario
                try:
                    user_answer = int(user_answer)
                    expected = 30
                    
                    if user_answer == expected:
                        add_result(1, "Cálculo", "¿Cuantos ID tiene la base de datos?", "Correcto",
                                   f"Obtenido: {user_answer}")
                    else:
                        add_result(1, "Cálculo", "¿Cuantos ID tiene la base de datos?", "Incorrecto",
                                  f"Esperado: 30, Obtenido: {user_answer}")
                except ValueError:
                    add_result(1, "Cálculo", "¿Cuantos ID tiene la base de datos?", "Error", 
                              f"Respuesta no es un número: {user_answer}")
        except Exception as e:
            add_result(1, "Cálculo", "¿Cuantos ID tiene la base de datos?", "Error",
                      f"Error al evaluar respuesta: {str(e)}")

        # Pregunta 2: Cambia el nombre de la columna "Seguimiento" por "Sentimiento"
        
        expected_col_name = "Sentimiento"
        actual_col_name = ws_user[f'{COLUMN_MAPPING["Sentimiento"]}{HEADER_ROW}'].value
        if actual_col_name == expected_col_name:
            add_result(2, "Edición y formato", "Cambia el nombre de la columna 'Seguimiento' por 'Sentimiento'",
                       "Correcto")
        else:
            add_result(2, "Edición y formato", "Cambia el nombre de la columna 'Seguimiento' por 'Sentimiento'",
                       "Incorrecto",
                       f"Nombre de columna en usuario: '{actual_col_name}', Esperado: '{expected_col_name}'")

        # Pregunta 3: Centra el contenido de todas las celdas verticalmente
        # Verificar el rango C5:K36 (excluyendo C36:I36)
        try:
            # Definir el rango específico
            start_row = 5  # C5
            end_row = 36   # K36
            
            # Verificar cada celda en el rango, excluyendo C36:I36
            all_cells_centered = True
            for row in range(start_row, end_row + 1):
                for col_letter in COLUMN_MAPPING.values():
                    # Saltar el rango C36:I36
                    if row == 36 and col_letter in [chr(i) for i in range(ord('C'), ord('I') + 1)]:
                        continue
                        
                    cell = ws_user[f"{col_letter}{row}"]
                    # Verificar el centrado horizontal y vertical
                    if cell.alignment.horizontal != 'center' or cell.alignment.vertical != 'center':
                        all_cells_centered = False
                        break
                if not all_cells_centered:
                    break

            if all_cells_centered:
                add_result(3, "Edición y formato", "Centrar el contenido de todas las celdas", "Correcto",
                           "Todas las celdas de la tabla están centradas.")
            else:
                add_result(3, "Edición y formato", "Centrar el contenido de todas las celdas", "Incorrecto",
                           "Alguna celda de la tabla no está centrada correctamente.")
        except Exception as e:
            add_result(3, "Edición y formato", "Centrar el contenido de todas las celdas", "Error",
                       f"Error al verificar el centrado: {str(e)}")

        # Pregunta 4: Ajusta el ancho de las columnas
        # Comparamos el ancho de las columnas relevantes (C a K)
        width_adjusted = True
        for col_letter in COLUMN_MAPPING.values():
            user_col_width = ws_user.column_dimensions[col_letter].width
            expected_col_width = ws_expected.column_dimensions[col_letter].width

            # Si el usuario no ha tocado el ancho, openpyxl puede devolver None
            # y el ancho por defecto es 8.43. Si es menor que un mínimo razonable, es incorrecto.
            if user_col_width is None or user_col_width < 8:  # Un ancho mínimo para que sea legible
                width_adjusted = False
                break
            # Además, podemos comparar con el ancho del archivo esperado, si lo definimos como 'óptimo'
            if expected_col_width is not None and user_col_width < expected_col_width * 0.9:  # 10% de tolerancia
                width_adjusted = False
                break

        if width_adjusted:
            add_result(4, "Edición y formato", "Ajusta el ancho de las columnas", "Correcto",
                       "Comparado con anchos de columnas esperados (con tolerancia) en el rango de la tabla.")
        else:
            add_result(4, "Edición y formato", "Ajusta el ancho de las columnas", "Incorrecto",
                       "El ancho de algunas columnas no parece ajustado correctamente en el rango de la tabla.")

        # Pregunta 5: Calcula el número total de llamadas registradas
        try:
            # Buscar la respuesta en M10
            answer_cell = "M10"  # Celda específica para la pregunta 5
            user_answer = ws_user[answer_cell].value
            
            if user_answer is None:
                add_result(5, "Fórmulas", "Calcula el número total de llamadas registradas", "Error", "Respuesta no encontrada")
            else:
                # Convertir a número si es necesario
                try:
                    user_answer = float(user_answer)
                    expected = 30.0  # Convertir a float para comparación consistente
                    
                    if user_answer == expected:
                        add_result(5, "Fórmulas", "Calcula el número total de llamadas registradas", "Correcto")
                    else:
                        add_result(5, "Fórmulas", "Calcula el número total de llamadas registradas", "Incorrecto",
                                  f"Esperado: 30, Obtenido: {user_answer}")
                except ValueError:
                    add_result(5, "Fórmulas", "Calcula el número total de llamadas registradas", "Error", 
                              f"Respuesta no es un número: {user_answer}")
        except Exception as e:
            add_result(5, "Fórmulas", "Calcula el número total de llamadas registradas", "Error",
                      f"Error al evaluar respuesta: {str(e)}")

        # Pregunta 6: Utiliza la función "Dar formato como tabla"
        try:
            # Verificar si hay formato de tabla aplicado
            table_found = False
            table_range_str = "C5:K35"  # Rango específico
            
            # Buscar tabla que cubra el rango C5:K35
            if ws_user.tables:
                for table_name, table_obj in ws_user.tables.items():
                    if table_obj.ref == table_range_str:
                        table_found = True
                        break
            
            # Si no hay tabla, verificar si hay filtros aplicados (para LibreOffice Calc)
            if not table_found:
                filter_applied = False
                for col_letter in COLUMN_MAPPING.values():
                    if ws_user.auto_filter.ref and ws_user.auto_filter.ref.startswith(col_letter):
                        filter_applied = True
                        break
            
            if table_found or filter_applied:
                add_result(6, "Fórmulas", "Utiliza la función 'Dar formato como tabla'", "Correcto",
                           "Formato de tabla o filtros aplicados correctamente en el rango C5:K35")
            else:
                add_result(6, "Fórmulas", "Utiliza la función 'Dar formato como tabla'", "Incorrecto",
                           "No se encontró formato de tabla ni filtros aplicados en el rango C5:K35")
        except Exception as e:
            add_result(6, "Fórmulas", "Utiliza la función 'Dar formato como tabla'", "Error",
                       f"Error al verificar el formato de tabla: {str(e)}")

        # Pregunta 7: Cuántas llamadas tuvieron un Sentimiento "Very Positive"
        try:
            # Buscar la respuesta en M12
            answer_cell = "M12"  # Celda específica para la pregunta 7
            user_answer = ws_user[answer_cell].value
            
            if user_answer is None:
                add_result(7, "Fórmulas", "Cuántas llamadas tuvieron un Sentimiento 'Very Positive'", "Error", "Respuesta no encontrada")
            else:
                # Convertir a número si es necesario
                try:
                    user_answer = int(user_answer)
                    expected = 3  # Convertir a float para comparación consistente
                    
                    if user_answer == expected:
                        add_result(7, "Fórmulas", "Cuántas llamadas tuvieron un Sentimiento 'Very Positive'", "Correcto",
                                   f"Esperado: 3, Obtenido: {user_answer}")
                    else:
                        add_result(7, "Fórmulas", "Cuántas llamadas tuvieron un Sentimiento 'Very Positive'", "Incorrecto",
                                   f"Esperado: 3, Obtenido: {user_answer}")
                except ValueError:
                    add_result(7, "Fórmulas", "Cuántas llamadas tuvieron un Sentimiento 'Very Positive'", "Error", 
                              f"Respuesta no es un número: {user_answer}")
        except Exception as e:
            add_result(7, "Fórmulas", "Cuántas llamadas tuvieron un Sentimiento 'Very Positive'", "Error",
                      f"Error al evaluar respuesta: {str(e)}")

        # Pregunta 8: Calcula la duración promedio de las llamadas
        try:
            # Verificar la fórmula en K36
            formula_cell = "K36"
            formula = ws_user[formula_cell].value
            
            # Verificar si la fórmula es correcta (acepta tanto PROMEDIO como AVERAGE)
            formula_correct = False
            
            # Verificar si la fórmula es correcta (considerando mayúsculas/minúsculas)
            if formula:
                formula_upper = str(formula).upper()
                if formula_upper == "=PROMEDIO(K6:K35)" or formula_upper == "=AVERAGE(K6:K35)":
                    formula_correct = True
            
            # Verificar la respuesta en M13
            answer_cell = "M13"
            user_answer = ws_user[answer_cell].value
            
            if user_answer is None:
                add_result(8, "Fórmulas", "Calcula la duración promedio de las llamadas", "Error", "Respuesta no encontrada")
            else:
                # Convertir a número si es necesario
                try:
                    user_answer = int(user_answer)  # Convertir a entero
                    expected_answer = 27  # Valor redondeado al entero más cercano
                    
                    if formula_correct and user_answer == expected_answer:
                        add_result(8, "Fórmulas", "Calcula la duración promedio de las llamadas", "Correcto",
                                  "Fórmula y respuesta correctas")
                    elif not formula_correct:
                        add_result(8, "Fórmulas", "Calcula la duración promedio de las llamadas", "Incorrecto",
                                  f"Fórmula incorrecta: {formula}")
                    else:
                        add_result(8, "Fórmulas", "Calcula la duración promedio de las llamadas", "Incorrecto",
                                  f"Respuesta incorrecta: {user_answer}, esperado: 27")
                except ValueError:
                    add_result(8, "Fórmulas", "Calcula la duración promedio de las llamadas", "Error", 
                              f"Respuesta no es un número: {user_answer}")
        except Exception as e:
            add_result(8, "Fórmulas", "Calcula la duración promedio de las llamadas", "Error",
                      f"Error al evaluar respuesta: {str(e)}")

        # Pregunta 9: Ajusta el formato fecha a "dd/mm/yyyy"
        try:
            # Verificar el formato de fecha en el rango G6:G35
            date_format_correct = True
            start_row = 6  # G6
            end_row = 35   # G35
            
            # Verificar cada celda en el rango
            for row in range(start_row, end_row + 1):
                cell = ws_user[f'G{row}']
                if cell.data_type != 'd' or cell.number_format != 'dd/mm/yyyy':
                    date_format_correct = False
                    break

            if date_format_correct:
                add_result(9, "Fórmulas", "Ajusta el formato fecha a 'dd/mm/yyyy'", "Correcto",
                           "Formato de fecha correcto en todas las celdas del rango G6:G35")
            else:
                add_result(9, "Fórmulas", "Ajusta el formato fecha a 'dd/mm/yyyy'", "Incorrecto",
                           "El formato de fecha no es 'dd/mm/yyyy' o el tipo de dato no es fecha en alguna celda del rango G6:G35")
        except Exception as e:
            add_result(9, "Fórmulas", "Ajusta el formato fecha a 'dd/mm/yyyy'", "Error",
                      f"Error al verificar el formato de fecha: {str(e)}")

        # Pregunta 10: Ordena la tabla por 'Puntuación' de mayor a menor y dime ¿Cuál es el puntaje máximo?
        try:
            # Verificar ordenación en la columna F (Puntuación)
            order_correct = True
            score_col = "F"  # Columna de Puntuación
            
            # Verificar que los valores estén ordenados de mayor a menor
            for row in range(6, 35):  # De F6 a F35
                current_cell = ws_user[f'{score_col}{row}']
                next_cell = ws_user[f'{score_col}{row + 1}']
                
                if current_cell.value is not None and next_cell.value is not None:
                    if current_cell.value < next_cell.value:
                        order_correct = False
                        break
            
            # Verificar la respuesta en M15
            answer_cell = "M15"
            user_answer = ws_user[answer_cell].value
            
            if user_answer is None:
                add_result(10, "Fórmulas", "Ordena la tabla por 'Puntuación' de mayor a menor y puntaje máximo",
                           "Error", "Respuesta no encontrada")
            else:
                try:
                    user_answer = int(user_answer)  # Convertir a entero
                    expected_answer = 10  # Valor máximo esperado
                    
                    if order_correct and user_answer == expected_answer:
                        add_result(10, "Fórmulas", "Ordena la tabla por 'Puntuación' de mayor a menor y puntaje máximo",
                                  "Correcto", "Ordenación y respuesta correctas")
                    elif not order_correct:
                        add_result(10, "Fórmulas", "Ordena la tabla por 'Puntuación' de mayor a menor y puntaje máximo",
                                  "Incorrecto", "Los valores no están ordenados de mayor a menor")
                    else:
                        add_result(10, "Fórmulas", "Ordena la tabla por 'Puntuación' de mayor a menor y puntaje máximo",
                                  "Incorrecto", f"Respuesta incorrecta: {user_answer}, esperado: 10")
                except ValueError:
                    add_result(10, "Fórmulas", "Ordena la tabla por 'Puntuación' de mayor a menor y puntaje máximo",
                              "Error", f"Respuesta no es un número: {user_answer}")
        except Exception as e:
            add_result(10, "Fórmulas", "Ordena la tabla por 'Puntuación' de mayor a menor y puntaje máximo",
                      "Error", f"Error al evaluar respuesta: {str(e)}")

        # Pregunta 11: Cuantas llamadas hay con ese puntaje Máximo
        try:
            # Verificar la respuesta en M16
            answer_cell = "M16"
            user_answer = ws_user[answer_cell].value
            
            if user_answer is None:
                add_result(11, "Fórmulas", "Cuantas llamadas hay con ese puntaje Máximo", "Error", "Respuesta no encontrada")
            else:
                try:
                    user_answer = int(user_answer)  # Convertir a entero
                    expected_answer = 2  # Valor esperado
                    
                    if user_answer == expected_answer:
                        add_result(11, "Fórmulas", "Cuantas llamadas hay con ese puntaje Máximo", "Correcto")
                    else:
                        add_result(11, "Fórmulas", "Cuantas llamadas hay con ese puntaje Máximo", "Incorrecto",
                                  f"Respuesta incorrecta: {user_answer}, esperado: 2")
                except ValueError:
                    add_result(11, "Fórmulas", "Cuantas llamadas hay con ese puntaje Máximo", "Error", 
                              f"Respuesta no es un número: {user_answer}")
        except Exception as e:
            add_result(11, "Fórmulas", "Cuantas llamadas hay con ese puntaje Máximo", "Error",
                      f"Error al evaluar respuesta: {str(e)}")

        # Pregunta 12: Si el "ID" de un cliente es PJL-11752230. Dime cual es el nombre y apellido al que corresponde
        try:
            # Verificar la respuesta en M17
            answer_cell = "M17"
            user_answer = ws_user[answer_cell].value
            
            if user_answer is None:
                add_result(12, "Fórmulas", "Si el 'ID' de un cliente es PJL-11752230. Dime cual es el nombre y apellido al que corresponde", "Error", "Respuesta no encontrada")
            else:
                # Convertir a minúsculas para comparación insensible a mayúsculas/minúsculas
                user_answer_lower = str(user_answer).lower()
                expected_name = "linda lopez"
                
                if user_answer_lower == expected_name:
                    add_result(12, "Fórmulas", "Si el 'ID' de un cliente es PJL-11752230. Dime cual es el nombre y apellido al que corresponde", "Correcto")
                else:
                    add_result(12, "Fórmulas", "Si el 'ID' de un cliente es PJL-11752230. Dime cual es el nombre y apellido al que corresponde", "Incorrecto",
                              f"Respuesta incorrecta: {user_answer}, esperado: 'Linda Lopez'")
        except Exception as e:
            add_result(12, "Fórmulas", "Si el 'ID' de un cliente es PJL-11752230. Dime cual es el nombre y apellido al que corresponde", "Error",
                      f"Error al evaluar respuesta: {str(e)}")

        # Pregunta 13: Resalta en Rojo las celdas de la columna "Puntuación" que sean inferiores a 5 (<5)
        try:
            # Verificar formato condicional en el rango F6:F35
            score_col = "F"  # Columna de Puntuación
            format_range = f"{score_col}6:{score_col}35"
            
            # Verificar si hay formato condicional correcto
            conditional_formatting_correct = False
            
            # Verificar cada celda en el rango
            for row in range(6, 36):  # F6:F35
                cell = ws_user[f'{score_col}{row}']
                if cell.value is not None and cell.value < 5:
                    # Verificar si la celda está en rojo
                    if cell.font and cell.font.color and cell.font.color.rgb == 'FFFF0000':
                        conditional_formatting_correct = True
                    else:
                        conditional_formatting_correct = False
                        break
            
            if conditional_formatting_correct:
                add_result(13, "Fórmulas", "Resalta en Rojo las celdas de la columna 'Puntuación' que sean inferiores a 5 (<5)",
                           "Correcto", "Formato condicional correcto en todas las celdas menores a 5")
            else:
                add_result(13, "Fórmulas", "Resalta en Rojo las celdas de la columna 'Puntuación' que sean inferiores a 5 (<5)",
                           "Incorrecto", "No todas las celdas menores a 5 están en rojo o hay celdas con formato incorrecto")
        except Exception as e:
            add_result(13, "Fórmulas", "Resalta en Rojo las celdas de la columna 'Puntuación' que sean inferiores a 5 (<5)",
                      "Error", f"Error al evaluar formato condicional: {str(e)}")

        # Pregunta 14: Crea un gráfico de barras para mostrar la relación entre "Nombre del cliente" y "Puntuación".
        chart1_found = False
        for sheet_name in wb_user.sheetnames:
            ws_current = wb_user[sheet_name]
            for chart_obj in ws_current._charts:
                if chart_obj.type == 'bar':
                    chart1_found = True
                    break
            if chart1_found: break
        if chart1_found:
            add_result(14, "Gráficos", "Crea un gráfico de barras (Nombre del cliente vs Puntuación)", "Correcto",
                       "Gráfico de barras detectado (verificación superficial de tipo).")
        else:
            add_result(14, "Gráficos", "Crea un gráfico de barras (Nombre del cliente vs Puntuación)", "Incorrecto",
                       "No se encontró un gráfico de barras adecuado.")

        # Pregunta 15: Crea un gráfico que permita mostrar la relación entre la "Duración de la Llamada" y la "Puntuación".
        chart2_found = False
        for sheet_name in wb_user.sheetnames:
            ws_current = wb_user[sheet_name]
            for chart_obj in ws_current._charts:
                if chart_obj.type in ['scatter', 'bar', 'line']:
                    chart2_found = True
                    break
            if chart2_found: break
        if chart2_found:
            add_result(15, "Gráficos", "Crea un gráfico (Duración de la Llamada vs Puntuación)", "Correcto",
                       "Gráfico de relación Duración/Puntuación detectado (verificación superficial de tipo).")
        else:
            add_result(15, "Gráficos", "Crea un gráfico (Duración de la Llamada vs Puntuación)", "Incorrecto",
                       "No se encontró un gráfico adecuado para la relación Duración/Puntuación.")

        wb_user.close()
        wb_expected.close()

    except FileNotFoundError:
        add_result("", "", f"Error: Archivo de usuario no encontrado: {user_filename}", "Error",
                   "Asegúrate de que el archivo exista en la carpeta 'user_submissions'.")
    except KeyError as e:
        add_result("", "", f"Error en la evaluación de {user_filename}", "Error",
                   f"Hoja 'Datos' no encontrada o columna faltante: {e}. Asegúrate que la hoja se llama 'Datos' y las columnas son correctas.")
    except Exception as e:
        add_result("", "", f"Error inesperado al evaluar {user_filename}", "Error", f"Detalles: {e}")

    add_result("", "", "--- Fin de evaluación ---", "")


def generate_report():
    """Genera el archivo Excel con los resultados de la evaluación."""
    df_results = pd.DataFrame(results)

    # Crear un nuevo workbook y sheet para los resultados
    wb_results = Workbook()
    ws_report = wb_results.active
    ws_report.title = "Resultados_Evaluacion"

    # Escribir el DataFrame en la hoja
    for r_idx, row in enumerate(dataframe_to_rows(df_results, index=False, header=True), 1):
        ws_report.append(row)

    # Aplicar formato a los encabezados
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    for cell in ws_report[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Ajustar ancho de columnas
    for col in ws_report.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value is not None and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        if adjusted_width > 100: adjusted_width = 100
        ws_report.column_dimensions[column].width = adjusted_width

    # Formato condicional para el estado
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

    for row_idx in range(2, ws_report.max_row + 1):
        status_cell = ws_report.cell(row=row_idx, column=4)
        if status_cell.value == "Incorrecto":
            for col_idx in range(1, ws_report.max_column + 1):
                ws_report.cell(row=row_idx, column=col_idx).fill = red_fill
        elif status_cell.value == "Correcto":
            for col_idx in range(1, ws_report.max_column + 1):
                ws_report.cell(row=row_idx, column=col_idx).fill = green_fill
        elif status_cell.value and str(status_cell.value).startswith("---"):
            for col_idx in range(1, ws_report.max_column + 1):
                ws_report.cell(row=row_idx, column=col_idx).fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0",
                                                                               fill_type="solid")
                ws_report.cell(row=row_idx, column=col_idx).font = Font(bold=True)

    wb_results.save(RESULTS_FILE)
    print(f"\nInforme de evaluación generado en '{RESULTS_FILE}'")


def main():
    """Función principal para ejecutar la evaluación."""
    print("Iniciando la evaluación de archivos de usuario...")
    if not os.path.exists(SUBMISSIONS_DIR):
        print(
            f"Error: No se encontró la carpeta de envíos '{SUBMISSIONS_DIR}'. Crea esta carpeta y coloca los archivos de los usuarios aquí.")
        return

    submission_files = [f for f in os.listdir(SUBMISSIONS_DIR) if f.endswith('.xlsx') or f.endswith('.xlsm')]

    if not submission_files:
        print(f"No se encontraron archivos .xlsx o .xlsm en '{SUBMISSIONS_DIR}'.")
        return

    for filename in submission_files:
        submission_path = os.path.join(SUBMISSIONS_DIR, filename)
        print(f"Procesando: {filename}")
        evaluate_submission(submission_path)

    generate_report()
    print("Evaluación completada.")


if __name__ == "__main__":
    main()